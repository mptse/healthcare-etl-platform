import os
import csv
import tempfile
from datetime import datetime

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, StreamingHttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated

from apps.analytics.decorators import role_required
from apps.etl.models import ETLLog, RegistroClinico
from .processor import run_etl


# ── API Views ────────────────────────────────────────────────────────

class EjecutarETLView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        archivo = request.FILES.get('archivo')

        if archivo:
            sufijo = '.xlsx' if archivo.name.endswith('.xlsx') else '.csv'
            with tempfile.NamedTemporaryFile(delete=False, suffix=sufijo) as tmp:
                for chunk in archivo.chunks():
                    tmp.write(chunk)
                ruta = tmp.name
            nombre_archivo = archivo.name
        else:
            ruta = os.path.join(
                settings.BASE_DIR, 'backend', 'apps', 'etl', 'data', 'dataset_clinico.xlsx'
            )
            nombre_archivo = 'dataset_clinico.xlsx'

        if not os.path.exists(ruta):
            return Response(
                {'error': 'Archivo no encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )

        log = ETLLog.objects.create(
            usuario=request.user,
            estado='en_proceso',
            archivo_fuente=nombre_archivo,
        )

        resultado = run_etl(ruta)

        log.registros_procesados = resultado.get('registros_procesados', 0)
        log.registros_fallidos = resultado.get('registros_fallidos', 0)
        log.tiempo_ejecucion = resultado.get('tiempo_ejecucion', 0)
        log.estado = 'exitoso' if resultado.get('exito') else 'fallido'
        log.log_detalle = resultado.get('log_detalle', '')
        log.mensaje_error = resultado.get('mensaje_error', '')
        log.save()

        if resultado.get('exito'):
            return Response({
                'mensaje': 'ETL ejecutado correctamente.',
                'registros_procesados': log.registros_procesados,
                'log_id': log.id,
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'error': 'El ETL falló.',
                'detalle': log.mensaje_error,
                'log_id': log.id,
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HistorialETLView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        logs = ETLLog.objects.select_related('usuario').all()[:50]
        data = [{
            'id': log.id,
            'fecha_ejecucion': log.fecha_ejecucion.strftime('%Y-%m-%d %H:%M:%S'),
            'usuario': log.usuario.username if log.usuario else 'Sistema',
            'archivo_fuente': log.archivo_fuente,
            'registros_procesados': log.registros_procesados,
            'registros_fallidos': log.registros_fallidos,
            'tiempo_ejecucion': log.tiempo_ejecucion,
            'estado': log.estado,
        } for log in logs]
        return Response({'historial': data, 'total': len(data)})


class DetalleETLView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            log = ETLLog.objects.get(pk=pk)
        except ETLLog.DoesNotExist:
            return Response({'error': 'Log no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        return Response({
            'id': log.id,
            'fecha_ejecucion': log.fecha_ejecucion.strftime('%Y-%m-%d %H:%M:%S'),
            'usuario': log.usuario.username if log.usuario else 'Sistema',
            'archivo_fuente': log.archivo_fuente,
            'registros_procesados': log.registros_procesados,
            'registros_fallidos': log.registros_fallidos,
            'tiempo_ejecucion': log.tiempo_ejecucion,
            'estado': log.estado,
            'log_detalle': log.log_detalle,
            'mensaje_error': log.mensaje_error,
        })


# ── Vistas web ───────────────────────────────────────────────────────

@login_required
@role_required(allowed_roles=['Analista', 'Admin'])
def historial_web(request):
    logs = ETLLog.objects.select_related('usuario').all()
    return render(request, 'etl/historial_etl.html', {'historial': logs})


@login_required
@role_required(allowed_roles=['Analista', 'Admin'])
def subir_csv(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')

        if not archivo:
            messages.error(request, 'No se seleccionó ningún archivo.')
            return redirect('subir_csv')

        if not archivo.name.endswith(('.csv', '.xlsx')):
            messages.error(request, 'Solo se permiten archivos CSV o Excel (.xlsx).')
            return redirect('subir_csv')

        sufijo = '.xlsx' if archivo.name.endswith('.xlsx') else '.csv'
        with tempfile.NamedTemporaryFile(delete=False, suffix=sufijo) as tmp:
            for chunk in archivo.chunks():
                tmp.write(chunk)
            ruta = tmp.name

        log = ETLLog.objects.create(
            usuario=request.user,
            estado='en_proceso',
            archivo_fuente=archivo.name,
        )

        resultado = run_etl(ruta)  # ← retorna dict

        log.registros_procesados = resultado.get('registros_procesados', 0)
        log.registros_fallidos   = resultado.get('registros_fallidos', 0)
        log.tiempo_ejecucion     = resultado.get('tiempo_ejecucion', 0.0)
        log.log_detalle          = resultado.get('log_detalle', '')
        log.mensaje_error        = resultado.get('mensaje_error', '')
        log.estado               = 'exitoso' if resultado.get('exito') else 'fallido'
        log.save()

        if resultado.get('exito'):
            try:
                from apps.ml.trainer import entrenar_modelo
                entrenar_modelo()
            except Exception:
                pass
            messages.success(
                request,
                f'✓ "{archivo.name}" procesado: '
                f'{log.registros_procesados} registros en {log.tiempo_ejecucion}s.'
            )
        else:
            messages.error(
                request,
                f'Error al procesar "{archivo.name}". Revisa el historial.'
            )

        try:
            os.remove(ruta)
        except Exception:
            pass

        return redirect('historial_etl')

    return render(request, 'etl/subir_csv.html')


@login_required
@role_required(allowed_roles=['Analista', 'Admin'])
def eliminar_log(request, pk):
    from django.shortcuts import get_object_or_404
    log = get_object_or_404(ETLLog, pk=pk)
    if request.method == 'POST':
        log.delete()
        messages.success(request, f'Ejecución #{pk} eliminada.')
    return redirect('historial_etl')


@login_required
@role_required(allowed_roles=['Analista', 'Admin'])
def eliminar_todos_logs(request):
    if request.method == 'POST':
        total = ETLLog.objects.count()
        ETLLog.objects.all().delete()
        messages.success(request, f'{total} registros eliminados del historial.')
    return redirect('historial_etl')

# ── Exportación ──────────────────────────────────────────────────────

@login_required
@role_required(allowed_roles=['Analista', 'Admin', 'Médico'])
def exportar_csv(request):
    # Creamos un generador para enviar los datos por partes (streaming)
    def generar_csv():

        yield '\ufeff'
        # Usamos un objeto que Django pueda escribir
        class Echo:
            def write(self, value): return value

        writer = csv.writer(Echo())
        
        # Escribir cabecera
        yield writer.writerow([
            'ID', 'Nombres', 'Apellidos', 'Edad', 'Sexo',
            'Peso', 'Altura', 'IMC', 'Presión Sistólica', 'Presión Diastólica',
            'Frecuencia Cardíaca', 'Glucosa', 'Colesterol', 'Saturación O2',
            'Temperatura', 'Fumador', 'Consumo Alcohol', 'Actividad Física',
            'Diagnóstico', 'Riesgo', 'Fecha Consulta'
        ])

        # Usamos .iterator() para no cargar todo en RAM
        registros = RegistroClinico.objects.select_related('paciente').iterator()
        
        for r in registros:
            yield writer.writerow([
                r.paciente.identificacion, r.paciente.nombres, r.paciente.apellidos,
                r.paciente.edad, r.paciente.sexo,
                r.peso, r.altura, r.imc,
                r.presion_sistolica, r.presion_diastolica,
                r.frecuencia_cardiaca, r.glucosa, r.colesterol,
                r.saturacion_oxigeno, r.temperatura,
                'Sí' if r.fumador else 'No',
                'Sí' if r.consumo_alcohol else 'No',
                r.actividad_fisica, r.diagnostico_preliminar,
                r.riesgo_enfermedad, r.fecha_consulta,
            ])

    response = StreamingHttpResponse(generar_csv(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="pacientes_clinicos.csv"'
    return response

@login_required
@role_required(allowed_roles=['Analista', 'Admin', 'Médico'])
def exportar_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes Clínicos"

    headers = [
        'ID', 'Nombres', 'Apellidos', 'Edad', 'Sexo',
        'Peso', 'Altura', 'IMC', 'Presión Sistólica', 'Presión Diastólica',
        'Frecuencia Cardíaca', 'Glucosa', 'Colesterol', 'Saturación O2',
        'Temperatura', 'Fumador', 'Consumo Alcohol', 'Actividad Física',
        'Diagnóstico', 'Riesgo', 'Fecha Consulta'
    ]

    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    colores = {
        'Crítico': 'FFCCCC',
        'Alto': 'FFE5CC',
        'Medio': 'FFFACC',
        'Bajo': 'CCFFCC',
    }

    registros = RegistroClinico.objects.select_related('paciente').all()
    for row_num, r in enumerate(registros, 2):
        datos = [
            r.paciente.identificacion, r.paciente.nombres, r.paciente.apellidos,
            r.paciente.edad, r.paciente.sexo,
            r.peso, r.altura, r.imc,
            r.presion_sistolica, r.presion_diastolica,
            r.frecuencia_cardiaca, r.glucosa, r.colesterol,
            r.saturacion_oxigeno, r.temperatura,
            'Sí' if r.fumador else 'No',
            'Sí' if r.consumo_alcohol else 'No',
            r.actividad_fisica, r.diagnostico_preliminar,
            r.riesgo_enfermedad, str(r.fecha_consulta),
        ]
        color = colores.get(r.riesgo_enfermedad, 'FFFFFF')
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=row_num, column=col, value=valor)
            cell.fill = fill

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pacientes_clinicos.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required(allowed_roles=['Analista', 'Admin', 'Médico'])
def exportar_pdf(request):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1*cm
    )

    styles = getSampleStyleSheet()
    elementos = []

    # Título
    titulo_style = ParagraphStyle(
        'titulo', parent=styles['Title'],
        fontSize=16, textColor=colors.HexColor('#0d6efd'),
        spaceAfter=6
    )
    subtitulo_style = ParagraphStyle(
        'subtitulo', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey,
        spaceAfter=12
    )

    elementos.append(Paragraph("HealthAnalytics IPS — Reporte Clínico", titulo_style))
    elementos.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} por {request.user.username}",
        subtitulo_style
    ))
    elementos.append(Spacer(1, 0.3*cm))

    # Tabla
    headers = [
        'Paciente', 'Edad', 'Sexo', 'Glucosa',
        'Presión', 'Sat. O₂', 'IMC', 'Diagnóstico', 'Riesgo', 'Fecha'
    ]

    data = [headers]
    registros = RegistroClinico.objects.select_related('paciente').all()[:500]

    colores_riesgo = {
        'Crítico': colors.HexColor('#FFCCCC'),
        'Alto': colors.HexColor('#FFE5CC'),
        'Medio': colors.HexColor('#FFFACC'),
        'Bajo': colors.HexColor('#CCFFCC'),
    }

    estilos_filas = []
    for i, r in enumerate(registros, 1):
        fila = [
            f"{r.paciente.nombres} {r.paciente.apellidos}",
            str(r.paciente.edad),
            r.paciente.sexo,
            str(r.glucosa),
            f"{r.presion_sistolica}/{r.presion_diastolica}",
            f"{r.saturacion_oxigeno}%",
            str(round(r.imc, 1)),
            r.diagnostico_preliminar[:25] if r.diagnostico_preliminar else 'N/A',
            r.riesgo_enfermedad,
            str(r.fecha_consulta),
        ]
        data.append(fila)

        color = colores_riesgo.get(r.riesgo_enfermedad)
        if color:
            estilos_filas.append(('BACKGROUND', (0, i), (-1, i), color))

    tabla = Table(data, repeatRows=1)
    estilo_base = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]

    tabla.setStyle(TableStyle(estilo_base + estilos_filas))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_clinico.pdf"'
    return response